#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PATTERN SCANNER — Kaufpunkte nach Regelwerk (6 bullische Chartmuster)
=====================================================================
Liest eine Finviz-CSV ein, zieht Kurshistorie über die Twelve Data API
und berechnet für jeden Ticker bis zu 3 Kaufpunkte inkl. Strategie-
Kennzeichnung, Stop und Kursziel. Output: farbcodierte Excel-Tabelle.

Muster (gemäß Regelwerk):
  1. Darvas Box
  2. Minervini Trend Template (Filter, kein eigener Kaufpunkt)
  3. Volatility Contraction Pattern (VCP)
  4. Cup & Handle
  5. Rectangle Top
  6. High & Tight Flag

Aufruf:
  export TWELVE_DATA_API_KEY="dein_key"
  python pattern_scanner.py finviz.csv --out kaufpunkte.xlsx
  python pattern_scanner.py finviz.csv --out kaufpunkte.xlsx --ntfy mein-topic

Hinweise:
  - Free-Tier Twelve Data: 8 Calls/min → Script drosselt automatisch
    (--rate 8). 87 Ticker + SPY ≈ 11 Minuten Laufzeit.
  - Zwischenspeicher in ./.cache/ — bei Wiederholung am selben Tag
    werden keine API-Calls verbraucht.
  - RS-Rank wird als Perzentil INNERHALB der eingelesenen Liste plus
    Vergleich gegen SPY berechnet (Näherung an IBD-RS, siehe README).
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, date
from pathlib import Path

import numpy as np
import pandas as pd

try:
    import requests
except ImportError:
    sys.exit("Bitte installieren: pip install requests pandas numpy scipy openpyxl")

from scipy.signal import argrelextrema
from scipy.stats import linregress

import cup_handle_v2  # Cup & Handle auf Wochenbasis, "Giant Base"
import listen         # Wochenlisten: Vereinigung und Darvas-Recht
import exit_regeln    # Stop-Deckel und die eine Risiko-Formel
import ntfy_verlauf   # merkt sich jede verschickte Meldung fuer den Freitags-Putz
import positionen     # offene Positionen samt Exit-Regelwerk
import gewinnzonen_lauf  # Kapitel 12: Gewinnzonen je offener Beobachtung
import red_to_green   # Kapitel 9: Fokusliste fuer den Live-Waechter
import shakeout       # Kapitel 10: Spring samt Sekundaertest-Warteliste
import trigger_logbuch  # schreibt jedes Signal mit, gekauft oder nicht
import volumen        # IBD Volume % Change, Kurve je Aktie
import zahlen_termine  # Wer heute Abend berichtet
from config import CFG as ZENTRAL, hoechstens, mind_erreicht, pruefe_config

pruefe_config()       # faengt widerspruechliche Schwellwerte sofort ab

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------

API_BASE = "https://api.twelvedata.com/time_series"
OUTPUTSIZE = 420          # ~420 Handelstage: reicht für 252d-Rolling + Puffer
CACHE_DIR = Path(".cache")
BENCHMARK = "SPY"

# Regelwerk-Parameter. Die Werte, die sich der Scanner MIT dem Waechter
# teilt, kommen seit 28.07.2026 aus config.py — der einen Quelle der
# Wahrheit (Gerhards Aufraeumschritt 2). Alles, was nur den Scanner
# betrifft (Cup-Masse, VCP-Swings, Flaggen), bleibt hier stehen.
CFG = {
    # Darvas
    "darvas_lookback_52w": ZENTRAL["lookback"]["jahr_tage"],
    "darvas_box_days": ZENTRAL["darvas"]["box_tage"],
    # WAR 20 — der Waechter rechnete gleichzeitig mit einem anderen
    # Fenster. Genau diese stille Uneinheitlichkeit sollte weg: jetzt
    # einheitlich 10 Tage fuer Scanner UND Waechter.
    "darvas_vol_avg": ZENTRAL["volumen"]["fenster_tage"],
    # Minervini
    "tt_ma_slope_days": ZENTRAL["ma"]["kurz"],   # MA200[t] vs MA200[t-21]
    "tt_min_above_low": 0.25,        # mind. 25 % über 52W-Tief
    "tt_max_below_high": 0.25,       # max. 25 % unter 52W-Hoch
    "tt_rs_min": 70,                 # RS-Perzentil
    # VCP
    "vcp_swing_order": 5,            # Fenster für Swing-Erkennung
    "vcp_min_contractions": 2,
    "vcp_max_contractions": 6,
    "vcp_max_last_depth": 0.12,      # letzte Kontraktion idealerweise eng
    "vcp_vol_breakout": ZENTRAL["volumen"]["breakout_faktor_vcp"],
    "vcp_stop_pct": 0.08,            # Minervini: 7-8 % unter Einstieg
    # Cup & Handle
    "cup_min_len": 25,               # ~5 Wochen
    "cup_max_len": 130,              # ~6 Monate
    "cup_min_depth": 0.12,
    "cup_max_depth": 0.50,
    "cup_rim_tolerance": 0.06,       # Ränder auf ähnlichem Niveau (±6 %)
    "handle_max_len": 20,
    "handle_min_len": 4,
    "handle_max_retrace": 1 / 3,     # max 1/3 der Cup-Höhe
    "cup_min_score": 80,             # Muster muss zu mind. 80 % erfüllt sein
    # Rectangle
    "rect_lookback": 65,
    "rect_band": 0.02,               # Cluster-Toleranz ±2 %
    "rect_min_touches": 2,
    # Umsatzwachstum (Regelwerk: "Fundamentaldaten-Filter")
    # CAN SLIM (O'Neil) verlangt mindestens 25 % Wachstum im jüngsten Quartal
    # gegenüber dem Vorjahr; Minervini nennt ähnliche Größenordnungen.
    "umsatz_min_wachstum": 0.25,
    # High & Tight Flag
    "htf_min_rise": 0.90,
    "htf_max_pole_days": 42,
    "htf_min_low_price": 1.0,
    "htf_max_flag_cal_days": 35,
    "htf_max_flag_range": 0.25,      # Flag-Range < 25 % der Masthöhe
}


# ---------------------------------------------------------------------------
# Datenabruf (Twelve Data) mit Cache & Rate-Limit
# ---------------------------------------------------------------------------

class RateLimiter:
    def __init__(self, calls_per_min: int):
        self.interval = 60.0 / max(1, calls_per_min)
        self._last = 0.0

    def wait(self):
        delta = time.time() - self._last
        if delta < self.interval:
            time.sleep(self.interval - delta)
        self._last = time.time()


# Ergebnis des Yahoo-Sammelabrufs. Yahoo liefert ALLE Ticker in einem
# einzigen Aufruf (87 Aktien in rund 3 Sekunden), waehrend Twelve Data im
# Gratistarif nur 8 Abfragen je Minute erlaubt und damit rund 11 Minuten
# braucht. Twelve Data bleibt als Rueckfallebene erhalten.
_YAHOO_DATEN: dict = {}
_YAHOO_GELAUFEN = False


def lade_yahoo_sammelabruf(tickers: list[str]) -> int:
    """Holt die Historien aller Ticker in EINEM Abruf von Yahoo.

    Liefert die Anzahl erfolgreich geholter Aktien. Schlaegt der Abruf fehl
    oder fehlt yfinance, bleibt der Speicher leer und jeder Ticker geht
    einzeln ueber Twelve Data — das System laeuft dann langsamer weiter,
    aber es laeuft."""
    global _YAHOO_GELAUFEN
    _YAHOO_GELAUFEN = True
    try:
        import yfinance as yf
    except ImportError:
        print("  yfinance nicht verfügbar — weiche auf Twelve Data aus.")
        return 0

    print(f"  Sammelabruf über Yahoo für {len(tickers)} Aktien …")
    t0 = time.time()
    try:
        roh = yf.download(" ".join(tickers), period="2y", interval="1d",
                          group_by="ticker", progress=False,
                          auto_adjust=False, threads=True)
    except Exception as e:
        print(f"  Yahoo-Sammelabruf fehlgeschlagen ({str(e)[:60]}) — Twelve Data übernimmt.")
        return 0

    for t in tickers:
        try:
            df = roh[t] if len(tickers) > 1 else roh
            df = df.dropna(subset=["Open", "High", "Low", "Close", "Volume"]).copy()
            if df.empty:
                continue
            df = df.reset_index()
            # Auf das Format bringen, das der Rest des Programms erwartet
            df = df.rename(columns={"Date": "datetime", "Datetime": "datetime",
                                    "Open": "open", "High": "high", "Low": "low",
                                    "Close": "close", "Volume": "volume"})
            df["datetime"] = pd.to_datetime(df["datetime"]).dt.tz_localize(None)
            for spalte in ("open", "high", "low", "close", "volume"):
                df[spalte] = pd.to_numeric(df[spalte], errors="coerce")
            df = df[["datetime", "open", "high", "low", "close", "volume"]]
            df = df.dropna().sort_values("datetime").reset_index(drop=True)
            if len(df) >= 60:
                _YAHOO_DATEN[t] = df
        except Exception:
            continue

    print(f"  Yahoo lieferte {len(_YAHOO_DATEN)} von {len(tickers)} Aktien "
          f"in {time.time() - t0:.1f} Sekunden.")
    return len(_YAHOO_DATEN)


def hole_fundamentals(tickers: list[str]) -> dict:
    """Holt Umsatz- und Gewinnwachstum je Aktie.

    Das Regelwerk sieht einen Fundamentaldaten-Filter vor ("Umsatzwachstum")
    und nennt dafuer Finnhub. Dessen Gratistarif liefert fuer US-Aktien aber
    keine brauchbaren Fundamentaldaten mehr, waehrend Yahoo sie mitliefert -
    ohne Schluessel und ohne zweiten Anbieter. Geprueft an sechs Aktien:
    alle mit Umsatzwachstum.

    Rueckgabe: {ticker: {"umsatzwachstum": float|None,
                         "gewinnwachstum": float|None}}
    Faellt der Abruf aus, bleibt der Wert None - dann wird nicht gefiltert,
    statt Aktien faelschlich auszuschliessen."""
    try:
        import yfinance as yf
    except ImportError:
        print("  yfinance fehlt — Umsatzfilter übersprungen.")
        return {}

    out = {}
    t0 = time.time()
    for ticker in tickers:
        try:
            info = yf.Ticker(ticker).info
            out[ticker] = {
                "umsatzwachstum": info.get("revenueGrowth"),
                "gewinnwachstum": (info.get("earningsGrowth")
                                   or info.get("earningsQuarterlyGrowth")),
            }
        except Exception:
            out[ticker] = {"umsatzwachstum": None, "gewinnwachstum": None}

    mit_daten = sum(1 for v in out.values() if v["umsatzwachstum"] is not None)
    print(f"  Fundamentaldaten: {mit_daten} von {len(tickers)} Aktien "
          f"in {time.time() - t0:.1f} Sekunden.")
    return out


def yahoo_einzeln(ticker: str) -> pd.DataFrame | None:
    """Einen einzelnen Ticker von Yahoo holen.

    Fuer Aufrufer, die keinen Sammelabruf machen — vor allem die
    Streamlit-App, die immer nur eine Aktie auf einmal anzeigt."""
    try:
        import yfinance as yf
    except ImportError:
        return None
    try:
        df = yf.download(ticker, period="2y", interval="1d",
                         progress=False, auto_adjust=False)
    except Exception:
        return None
    if df is None or df.empty:
        return None
    # Bei einzelnem Ticker liefert yfinance je nach Fassung mehrstufige Spalten
    if hasattr(df.columns, "levels"):
        df.columns = df.columns.droplevel(1)
    df = df.dropna(subset=["Open", "High", "Low", "Close", "Volume"]).reset_index()
    df = df.rename(columns={"Date": "datetime", "Datetime": "datetime",
                            "Open": "open", "High": "high", "Low": "low",
                            "Close": "close", "Volume": "volume"})
    try:
        df["datetime"] = pd.to_datetime(df["datetime"]).dt.tz_localize(None)
    except Exception:
        df["datetime"] = pd.to_datetime(df["datetime"])
    for spalte in ("open", "high", "low", "close", "volume"):
        df[spalte] = pd.to_numeric(df[spalte], errors="coerce")
    df = df[["datetime", "open", "high", "low", "close", "volume"]]
    df = df.dropna().sort_values("datetime").reset_index(drop=True)
    return df if len(df) >= 60 else None


def fetch_history(ticker: str, api_key: str, limiter: RateLimiter) -> pd.DataFrame | None:
    """Tageskurse (OHLCV) holen — mit Tages-Cache, damit Reruns gratis sind.

    Reihenfolge: Tages-Cache, dann Yahoo-Sammelabruf, dann Twelve Data."""
    CACHE_DIR.mkdir(exist_ok=True)
    cache_file = CACHE_DIR / f"{ticker}_{date.today().isoformat()}.csv"
    if cache_file.exists():
        c = pd.read_csv(cache_file); c["datetime"] = pd.to_datetime(c["datetime"]); return c

    # Aus dem Sammelabruf bedienen, falls vorhanden
    if ticker in _YAHOO_DATEN:
        # Leere Kerzen (Datum ohne Werte) verwerfen - siehe
        # breakout_watcher, Befund vom 18.08.2026: NaN-Zeilen vergiften
        # sonst Indikatoren und Durchschnitte.
        df = _YAHOO_DATEN[ticker].dropna(subset=["close"])
        df.to_csv(cache_file, index=False)
        return df

    # Kein Sammelabruf gelaufen — etwa in der Streamlit-App, die immer nur
    # eine einzelne Aktie anzeigt. Dann diesen einen Ticker von Yahoo holen.
    df = yahoo_einzeln(ticker)
    if df is not None:
        df.to_csv(cache_file, index=False)
        return df

    # Rueckfallebene: einzeln ueber Twelve Data
    if not api_key:
        return None
    limiter.wait()
    params = {
        "symbol": ticker,
        "interval": "1day",
        "outputsize": OUTPUTSIZE,
        "apikey": api_key,
        "order": "asc",
    }
    try:
        r = requests.get(API_BASE, params=params, timeout=30)
        data = r.json()
    except Exception as e:
        print(f"  [{ticker}] Netzwerkfehler: {e}")
        return None

    if data.get("status") == "error" or "values" not in data:
        print(f"  [{ticker}] API-Fehler: {data.get('message', 'unbekannt')}")
        return None

    df = pd.DataFrame(data["values"])
    df["datetime"] = pd.to_datetime(df["datetime"])
    for col in ("open", "high", "low", "close", "volume"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna().sort_values("datetime").reset_index(drop=True)
    if len(df) < 60:
        print(f"  [{ticker}] Zu wenig Historie ({len(df)} Tage) — übersprungen")
        return None
    df.to_csv(cache_file, index=False)
    return df


def add_indicators(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    c = df["close"]
    df["ma21"] = c.rolling(21).mean()
    df["ma50"] = c.rolling(50).mean()
    df["ma150"] = c.rolling(150).mean()
    df["ma200"] = c.rolling(200).mean()
    # WAR fest auf 20 verdrahtet, obwohl config.py laengst ein anderes
    # Fenster vorgab — die Zahl wanderte von hier in die Box-Notiz und
    # widersprach damit dem, wogegen der Waechter tatsaechlich prueft.
    # Seit Gerhards Umbau vom 28.07.2026 zieht auch diese Stelle ihren
    # Massstab aus config.py (derzeit 50 Tage, IBD-Standard).
    df["vol_schnitt"] = df["volume"].rolling(CFG["darvas_vol_avg"]).mean()
    df["hi52"] = df["high"].rolling(CFG["darvas_lookback_52w"], min_periods=60).max()
    df["lo52"] = df["low"].rolling(CFG["darvas_lookback_52w"], min_periods=60).min()
    return df


def rs_score(df: pd.DataFrame) -> float | None:
    """IBD-ähnlicher RS-Score: gewichtete Returns 3/6/9/12 Monate
    (Gewichtung 2:1:1:1 auf das jüngste Quartal)."""
    c = df["close"]
    if len(c) < 252:
        # so viel nehmen wie da ist, mit gleicher Logik
        if len(c) < 63:
            return None
    def ret(days):
        if len(c) <= days:
            return c.iloc[-1] / c.iloc[0] - 1
        return c.iloc[-1] / c.iloc[-1 - days] - 1
    return 2 * ret(63) + ret(126) + ret(189) + ret(252)


# ---------------------------------------------------------------------------
# Hilfsfunktionen: Swings
# ---------------------------------------------------------------------------

def swing_points(df: pd.DataFrame, order: int):
    """Lokale Swing-Hochs/-Tiefs (Index-Positionen) via argrelextrema."""
    highs = argrelextrema(df["high"].values, np.greater_equal, order=order)[0]
    lows = argrelextrema(df["low"].values, np.less_equal, order=order)[0]
    # doppelte Plateaus entschärfen
    highs = _dedupe(highs)
    lows = _dedupe(lows)
    return highs, lows


def _dedupe(idx: np.ndarray, min_gap: int = 3) -> list[int]:
    out = []
    for i in idx:
        if not out or i - out[-1] >= min_gap:
            out.append(int(i))
    return out


# ---------------------------------------------------------------------------
# 1) DARVAS BOX
# ---------------------------------------------------------------------------

def detect_darvas(df: pd.DataFrame) -> dict | None:
    """52W-Hoch → Box-Top aus 3 Tagen (Hoch-Tag + 2 Folgetage) → Box-Bottom
    aus den 3 Tagen danach → Bestätigung wenn Kurs ≥3 Tage in der Box bleibt.
    Kaufpunkt: Box-Top + 1 Cent (Volumen-Bestätigung als Hinweis).

    Deckel und Boden umfassen je 3 Tage („3+3") — Gerhards Vorgabe vom
    22.07.2026; vorher zählte der Deckel versehentlich 4 Tage (Hoch-Tag
    plus 3), wodurch der Boden einen Tag zu spät begann."""
    n = len(df)
    look = min(n, CFG["darvas_lookback_52w"])
    win = df.iloc[-look:]
    hi_pos = int(win["high"].idxmax())          # Position des 52W-Hochs
    bars_after = n - 1 - hi_pos
    bd = CFG["darvas_box_days"]
    if bars_after < 2 * bd - 1:
        return None  # Box noch nicht fertig ausgebildet (Deckel 3 + Boden 3,
                     # der Hoch-Tag selbst zählt zum Deckel)
    # Frische-Regel: bewusst NICHT im Regelwerk-Dokument, aber von Gerhard
    # am 22.07.2026 ausdrücklich bestätigt — nur Boxen frisch nach neuem
    # Hoch melden, keine monatealten toten Formationen.
    if bars_after > ZENTRAL["darvas"]["frische_max_tage"]:
        return None  # 52W-Hoch zu alt

    # Box-Top: höchstes Hoch aus Hoch-Tag + 2 Folgetagen (3 Tage)
    top_win = df.iloc[hi_pos: hi_pos + bd]
    box_top = float(top_win["high"].max())
    top_end = hi_pos + bd - 1
    # Box-Bottom: tiefstes Tief der 3 Tage nach Box-Top-Fixierung
    bot_win = df.iloc[top_end + 1: top_end + 1 + bd]
    if len(bot_win) < bd:
        return None
    box_bottom = float(bot_win["low"].min())

    # Bestätigung: seither in der Box geblieben (Schlusskurse)?
    since = df.iloc[top_end + 1:]
    inside = since[(since["close"] <= box_top) & (since["close"] >= box_bottom)]
    last_close = float(since["close"].iloc[-1])
    confirmed = (len(inside) >= bd
                 and len(inside) == len(since)          # KEIN Ausreißer aus der Box
                 and box_bottom <= last_close <= box_top)

    last = df.iloc[-1]
    if not confirmed:
        # Kurs schon ausgebrochen oder Box gerissen → kein frisches Setup
        return None

    return {
        "strategie": "Darvas Box",
        "kaufpunkt": round(box_top + 0.01, 2),
        "stop": round(box_bottom - 0.01, 2),
        "ziel": None,
        "status": "Box bestätigt — auf Breakout mit Volumen warten",
        "notiz": f"Box {box_bottom:.2f}–{box_top:.2f}; Breakout nur mit Volumen "
                 f"über Ø{CFG['darvas_vol_avg']}d "
                 f"({last['vol_schnitt']:,.0f}) gültig",
    }


# ---------------------------------------------------------------------------
# 2) MINERVINI TREND TEMPLATE (Filter)
# ---------------------------------------------------------------------------

def check_trend_template(df: pd.DataFrame, rs_percentile: float | None) -> tuple[bool, int, list[str]]:
    last = df.iloc[-1]
    if pd.isna(last["ma200"]):
        return False, 0, ["Zu wenig Historie für MA200"]
    slope_ok = False
    if len(df) > 200 + CFG["tt_ma_slope_days"]:
        slope_ok = last["ma200"] > df["ma200"].iloc[-1 - CFG["tt_ma_slope_days"]]
    checks = {
        "Kurs > MA150 & MA200": last["close"] > last["ma150"] and last["close"] > last["ma200"],
        "MA150 > MA200": last["ma150"] > last["ma200"],
        "MA200 steigt (≥1 Monat)": bool(slope_ok),
        "MA50 > MA150 & MA200": last["ma50"] > last["ma150"] and last["ma50"] > last["ma200"],
        "Kurs > MA50": last["close"] > last["ma50"],
        "≥25 % über 52W-Tief": last["close"] >= last["lo52"] * (1 + CFG["tt_min_above_low"]),
        "≤25 % unter 52W-Hoch": last["close"] >= last["hi52"] * (1 - CFG["tt_max_below_high"]),
        "RS-Rank ≥ 70": rs_percentile is not None and mind_erreicht(rs_percentile, CFG["tt_rs_min"]),
    }
    failed = [k for k, v in checks.items() if not v]
    return len(failed) == 0, sum(checks.values()), failed


# ---------------------------------------------------------------------------
# 3) VCP
# ---------------------------------------------------------------------------

def detect_vcp(df: pd.DataFrame, tt_pass: bool) -> dict | None:
    """Serie abnehmender Kontraktionen + Volume Dry-Up.
    Basisvoraussetzung laut Regelwerk: Trend Template erfüllt."""
    if not tt_pass:
        return None
    sub = df.iloc[-160:].reset_index(drop=True)  # letzte ~7 Monate
    highs, lows = swing_points(sub, CFG["vcp_swing_order"])
    if len(highs) < 2 or len(lows) < 2:
        return None

    # Kontraktionen: Swing-Hoch → nächstes Swing-Tief danach
    contractions = []
    for h in highs:
        nxt = [l for l in lows if l > h]
        if not nxt:
            continue
        l = nxt[0]
        depth = 1 - sub["low"].iloc[l] / sub["high"].iloc[h]
        if 0.005 < depth < 0.40:
            contractions.append({"h": h, "l": l, "depth": depth})

    if len(contractions) < CFG["vcp_min_contractions"]:
        return None
    contractions = contractions[-CFG["vcp_max_contractions"]:]

    depths = [c["depth"] for c in contractions]
    # monoton fallende Tiefen (T1 > T2 > T3 …) — kleine Toleranz von 10 %
    monotone = all(depths[i] > depths[i + 1] * 0.9 for i in range(len(depths) - 1))
    strictly = all(depths[i] > depths[i + 1] for i in range(len(depths) - 1))
    if not monotone:
        return None

    # Volume Dry-Up über die gesamte Formation
    start = contractions[0]["h"]
    vol_seg = sub["volume"].iloc[start:]
    slope = linregress(np.arange(len(vol_seg)), vol_seg.values).slope
    dryup = slope < 0

    pivot = float(sub["high"].iloc[contractions[-1]["h"]:].max())
    last_close = float(sub["close"].iloc[-1])
    if last_close > pivot * 1.02:
        return None  # schon > 2 % über Pivot — Zug abgefahren

    kp = round(pivot + 0.01, 2)
    seq = " → ".join(f"{d*100:.0f}%" for d in depths)
    return {
        "strategie": "VCP",
        "kaufpunkt": kp,
        "stop": round(kp * (1 - CFG["vcp_stop_pct"]), 2),
        "ziel": None,
        "status": ("VCP komplett" if (strictly and dryup)
                   else "VCP (Toleranz)" if dryup
                   else "VCP ohne sauberen Vol-Dry-Up"),
        "notiz": f"Kontraktionen: {seq}; Pivot {pivot:.2f}; "
                 f"Breakout braucht Vol ≥ {CFG['vcp_vol_breakout']*100:.0f}% vom Ø",
    }


# ---------------------------------------------------------------------------
# 4) CUP & HANDLE
# ---------------------------------------------------------------------------

def detect_cup_handle(df: pd.DataFrame) -> dict | None:
    """U-förmiger Cup (quadratischer Fit) + Handle im oberen Drittel.
    Liefert Toleranz-Score statt hartem Ja/Nein (Regelwerk Kap. 4)."""
    sub = df.iloc[-(CFG["cup_max_len"] + CFG["handle_max_len"] + 10):].reset_index(drop=True)
    n = len(sub)
    if n < CFG["cup_min_len"] + CFG["handle_min_len"]:
        return None

    highs, _ = swing_points(sub, 4)
    best = None
    for left in highs:
        if n - left < CFG["cup_min_len"]:
            continue
        left_high = float(sub["high"].iloc[left])
        seg = sub.iloc[left:]
        bot_rel = int(seg["low"].values.argmin())
        bottom = float(seg["low"].iloc[bot_rel])
        depth = 1 - bottom / left_high
        if not (mind_erreicht(depth, CFG["cup_min_depth"])
                and hoechstens(depth, CFG["cup_max_depth"])):
            continue
        # rechter Rand: erstes Wiedererreichen von ~linkem Rand nach dem Boden
        after_bot = seg.iloc[bot_rel:]
        reach = after_bot[after_bot["high"] >= left_high * (1 - CFG["cup_rim_tolerance"])]
        if reach.empty:
            continue
        right = int(reach.index[0])            # Position in sub
        cup_len = right - left
        if not (CFG["cup_min_len"] <= cup_len <= CFG["cup_max_len"]):
            continue
        # Symmetrie: Boden ungefähr mittig (25–75 % der Cup-Länge)
        bot_abs = left + bot_rel
        sym = (bot_abs - left) / cup_len
        if not (0.2 <= sym <= 0.8):
            continue
        # U-Form: quadratischer Fit über die Cup-Tiefs, Öffnung nach oben
        x = np.arange(cup_len + 1)
        y = sub["low"].iloc[left:right + 1].values
        coef = np.polyfit(x, y, 2)
        fit = np.polyval(coef, x)
        ss_res = float(np.sum((y - fit) ** 2))
        ss_tot = float(np.sum((y - y.mean()) ** 2)) or 1e-9
        r2 = 1 - ss_res / ss_tot
        u_ok = coef[0] > 0 and r2 > 0.55
        # Handle nach rechtem Rand
        handle = sub.iloc[right:]
        if len(handle) < CFG["handle_min_len"] or len(handle) > CFG["handle_max_len"] + 8:
            continue
        h_high = float(handle["high"].iloc[0:3].max())
        h_low = float(handle["low"].min())
        retrace = (h_high - h_low) / (left_high - bottom + 1e-9)
        in_upper_third = h_low >= bottom + (left_high - bottom) * (2 / 3)
        if not hoechstens(retrace, CFG["handle_max_retrace"]) or not in_upper_third:
            continue

        # VOLUMENVERLAUF — laut Regelwerk Pflichtbestandteil des Musters:
        # "Volumen fällt tendenziell im Verlauf des Cups (besonders am Boden)
        #  und steigt wieder Richtung rechtem Rand."
        # "Volumen sollte während des Handles niedrig/rückläufig sein."
        # Das wurde bisher gar nicht geprüft.
        cup_vol = sub["volume"].iloc[left:right + 1].values
        drittel = max(1, len(cup_vol) // 3)
        vol_links = float(cup_vol[:drittel].mean())
        vol_boden = float(cup_vol[drittel:2 * drittel].mean()) if len(cup_vol) > drittel else vol_links
        vol_rechts = float(cup_vol[-drittel:].mean())
        # Am Boden soll weniger gehandelt werden als am linken Rand
        vol_trocknet = vol_boden < vol_links
        # Richtung rechtem Rand soll es wieder anziehen
        vol_zieht_an = vol_rechts > vol_boden
        # Im Handle soll es ruhig sein — gemessen am Cup-Durchschnitt
        vol_handle = float(handle["volume"].mean())
        vol_cup_schnitt = float(cup_vol.mean()) or 1.0
        handle_ruhig = vol_handle < vol_cup_schnitt

        # Score: 100 Punkte, davon 25 für den Volumenverlauf. Ein Muster ohne
        # passendes Volumen ist nach Regelwerk kein sauberes Cup & Handle.
        vol_punkte = (10 * (1 if vol_trocknet else 0)
                      + 8 * (1 if vol_zieht_an else 0)
                      + 7 * (1 if handle_ruhig else 0))
        score = (
            30 * min(1.0, r2 / 0.85)
            + 15 * (1 - abs(sym - 0.5) * 2)
            + 15 * (1 - retrace / CFG["handle_max_retrace"])
            + 15 * (1 if u_ok else 0.3)
            + vol_punkte
        )
        cand = {"left_high": left_high, "bottom": bottom, "h_high": h_high,
                "depth": depth, "score": score, "cup_len": cup_len,
                "vol_trocknet": vol_trocknet, "vol_zieht_an": vol_zieht_an,
                "handle_ruhig": handle_ruhig}
        if best is None or cand["score"] > best["score"]:
            best = cand

    # Schwelle 80 von 100 (Vorgabe Mathias, 21.07.2026). Vorher stand hier
    # 55 — damit galten auch Formationen als Cup & Handle, die das Muster
    # nur knapp zur Hälfte erfüllten.
    if best is None or not mind_erreicht(best["score"], CFG["cup_min_score"]):
        return None
    kp = round(best["h_high"] + 0.01, 2)
    ziel = round(kp + (best["left_high"] - best["bottom"]), 2)
    return {
        "strategie": "Cup & Handle",
        "kaufpunkt": kp,
        # Die Punktzahl zusaetzlich als ZAHL. Sie stand bisher nur im
        # Text unter "status". Gerhards Einbau-Beispiel vom 04.08.2026
        # vergleicht die beiden Cup-Fassungen ueber ein Feld 'score' —
        # ohne dieses Feld haette die Tagesfassung immer gewonnen und
        # die Wochenfassung waere nie zum Zug gekommen. Die Rechnung
        # selbst ist unveraendert.
        "score": round(best["score"], 1),
        "stop": round(best["bottom"] + (best["left_high"] - best["bottom"]) * (2 / 3), 2),
        "ziel": ziel,
        "status": f"Score {best['score']:.0f}/100",
        "notiz": (f"Cup-Tiefe {best['depth']*100:.0f} %, Länge {best['cup_len']} Tage; "
                  f"Volumen: {'trocknet am Boden' if best['vol_trocknet'] else 'trocknet NICHT'}, "
                  f"{'zieht rechts an' if best['vol_zieht_an'] else 'zieht rechts nicht an'}, "
                  f"Handle {'ruhig' if best['handle_ruhig'] else 'unruhig'}; "
                  f"Ziel = Breakout + Cup-Höhe"),
    }


# ---------------------------------------------------------------------------
# 5) RECTANGLE TOP
# ---------------------------------------------------------------------------

def detect_rectangle(df: pd.DataFrame) -> dict | None:
    """Horizontale Range: ≥2 Berührungen oben UND unten.
    Kaufstopp 1 Cent über Rectangle-Top, Zusatzfilter Kurs > SMA21."""
    sub = df.iloc[-CFG["rect_lookback"]:].reset_index(drop=True)
    highs, lows = swing_points(sub, 3)
    if len(highs) < 2 or len(lows) < 2:
        return None

    hvals = sub["high"].iloc[highs].values
    lvals = sub["low"].iloc[lows].values
    top = float(np.median(hvals))
    bot = float(np.median(lvals))
    band = CFG["rect_band"]
    top_touch = int(np.sum(np.abs(hvals / top - 1) <= band))
    bot_touch = int(np.sum(np.abs(lvals / bot - 1) <= band))
    if top_touch < CFG["rect_min_touches"] or bot_touch < CFG["rect_min_touches"]:
        return None
    if (top - bot) / top < 0.03 or (top - bot) / top > 0.25:
        return None  # zu flach (Rauschen) oder zu breit (keine Range)

    last = df.iloc[-1]
    if float(last["close"]) > top * 1.02:
        return None  # bereits ausgebrochen
    above_sma21 = float(last["close"]) > float(last["ma21"]) if not pd.isna(last["ma21"]) else False

    kp = round(top + 0.01, 2)
    return {
        "strategie": "Rectangle Top",
        "kaufpunkt": kp,
        "stop": round(bot - 0.01, 2),
        "ziel": round(kp + (top - bot), 2),
        "status": ("Setup komplett (Kurs > SMA21)" if above_sma21
                   else "Range steht — SMA21-Filter noch NICHT erfüllt"),
        "notiz": f"Range {bot:.2f}–{top:.2f}; Berührungen oben {top_touch}, "
                 f"unten {bot_touch}; Ziel = Ausbruch + Rechteckhöhe",
    }


# ---------------------------------------------------------------------------
# 6) HIGH & TIGHT FLAG
# ---------------------------------------------------------------------------

def detect_htf(df: pd.DataFrame) -> dict | None:
    sub = df.iloc[-110:].reset_index(drop=True)
    n = len(sub)
    if n < 50:
        return None
    lows = sub["low"].values
    highs = sub["high"].values

    best = None
    for i in range(n - 10):
        lo = lows[i]
        if lo < CFG["htf_min_low_price"]:
            continue
        j_end = min(n, i + CFG["htf_max_pole_days"] + 1)
        seg = highs[i:j_end]
        j_rel = int(seg.argmax())
        hi = seg[j_rel]
        rise = hi / lo - 1
        if mind_erreicht(rise, CFG["htf_min_rise"]):
            cand = {"i": i, "j": i + j_rel, "lo": float(lo), "hi": float(hi), "rise": rise}
            if best is None or cand["rise"] > best["rise"]:
                best = cand
    if best is None:
        return None

    j = best["j"]
    flag = sub.iloc[j:]
    if len(flag) < 3:
        return None
    cal_days = (sub["datetime"].iloc[-1] - sub["datetime"].iloc[j]).days
    if cal_days > CFG["htf_max_flag_cal_days"]:
        return None
    pole_h = best["hi"] - best["lo"]
    flag_range = float(flag["high"].max() - flag["low"].min())
    if flag_range > pole_h * CFG["htf_max_flag_range"]:
        return None
    vol_slope = linregress(np.arange(len(flag)), flag["volume"].values).slope if len(flag) > 3 else -1
    flag_high = float(flag["high"].max())
    if float(sub["close"].iloc[-1]) > flag_high * 1.02:
        return None

    kp = round(flag_high + 0.01, 2)
    # BENOTUNG (Soreide-Ausbau, Gerhards Freigabe 31.08.2026): Leif
    # Soreide benotet jede Flagge und richtet die Positionsgroesse
    # danach. Messbar sind drei seiner Groessen: die ENGE der Flagge
    # (Spanne zur Masthoehe), die STEILHEIT des Masts (Prozent je
    # Handelstag) und die VOLUMEN-AUSTROCKNUNG in der Flagge. Die
    # Schwellen sind eine Erstkalibrierung und werden nach drei Monaten
    # Logbuch nachgemessen; RS bleibt bewusst draussen, die Note soll
    # allein aus dem Muster lesbar sein.
    pole_tage = max(1, j - best["i"])
    punkte = 0
    punkte += 2 if flag_range <= pole_h * 0.15 else (
        1 if flag_range <= pole_h * 0.20 else 0)
    steil = best["rise"] / pole_tage
    punkte += 2 if steil >= 0.03 else (1 if steil >= 0.022 else 0)
    punkte += 2 if vol_slope < 0 else 0
    note = "A" if punkte >= 5 else ("B" if punkte >= 3 else "C")
    return {
        "strategie": "High & Tight Flag",
        "kaufpunkt": kp,
        "stop": round(float(flag["low"].min()) - 0.01, 2),
        "ziel": None,
        # Gerhards Entscheid 22.07.2026: Volumenverlauf ist bei der HTF KEIN
        # Pflichtkriterium (sonst fiele das seltenste Muster oft ganz aus),
        # sondern wird im Status gekennzeichnet.
        "status": ("HTF komplett" if vol_slope < 0
                   else "HTF ohne Vol-Bestätigung") + f"; Note {note}",
        "notiz": f"Mast +{best['rise']*100:.0f}% in {j - best['i']} Tagen; "
                 f"Flag {cal_days} Kalendertage, Range {flag_range/pole_h*100:.0f}% der Masthöhe",
        # Interne Felder fuer den Innen-Einstieg und die Auswertung;
        # die Mappe liest nur ihre festen Spalten, hier stoert nichts.
        "flag_tage": int(len(flag)),
        "htf_note": note,
    }


def detect_htf_innen(df: pd.DataFrame) -> dict | None:
    """HTF Innen-Einstieg (Soreide-Ausbau, Gerhards Freigabe 31.08.2026).

    Leif Soreide (US-Meister 2019, Spezialist genau dieses Musters)
    kauft nicht erst den Riss des Flaggenhochs, sondern enge Stellen IN
    der Flagge: einen Inside Day (Hoch und Tief innerhalb des Vortags)
    oder den engsten Tag der Flagge. Kaufpunkt = Hoch dieses Tages plus
    1 Cent, Stop = sein Tief minus 1 Cent — das Risiko je Versuch wird
    drastisch kleiner, dafuer steigt die Fehlversuchsquote; Soreides
    Wort dazu: der beste Verlierer sein. Die Marke laeuft ZUSAETZLICH
    zur Flaggenhoch-Marke (Regelfrage G8, so entschieden — die Mappe
    traegt drei Kaufpunkte je Aktie, und der Waechter buendelt beide in
    EINE Meldung, wenn sie zugleich reissen).

    Gesucht wird nur in den JUENGSTEN fuenf Flaggentagen — eine alte
    enge Stelle ist keine Einstiegsmarke mehr. Ist das Kandidaten-Hoch
    schon ueberschritten oder liegt es nicht UNTER dem Flaggenhoch,
    gibt es nichts zu melden."""
    res = detect_htf(df)
    if not res or (res.get("flag_tage") or 0) < 3:
        return None
    flag = df.iloc[-int(res["flag_tage"]):]
    letzte = min(5, len(flag) - 1)
    kandidat = None
    for r in range(1, letzte + 1):
        tag = flag.iloc[-r]
        vortag = flag.iloc[-r - 1]
        if (float(tag["high"]) <= float(vortag["high"])
                and float(tag["low"]) >= float(vortag["low"])):
            kandidat, art = tag, "Inside Day"
            break
    if kandidat is None:
        fenster = flag.iloc[-letzte:]
        spannen = (fenster["high"] - fenster["low"]).astype(float)
        kandidat, art = fenster.loc[spannen.idxmin()], "engster Tag"
    kp = round(float(kandidat["high"]) + 0.01, 2)
    if kp >= res["kaufpunkt"]:
        return None
    if float(df["close"].iloc[-1]) > float(kandidat["high"]):
        return None
    return {
        "strategie": "HTF Innen-Einstieg",
        "kaufpunkt": kp,
        "stop": round(float(kandidat["low"]) - 0.01, 2),
        "ziel": None,
        "status": f"{art} in der Flagge; Note {res.get('htf_note', 'C')}",
        "notiz": res.get("notiz", ""),
        "htf_note": res.get("htf_note"),
    }


# ---------------------------------------------------------------------------
# Fallback-Kaufpunkte (wenn < 3 Muster aktiv)
# ---------------------------------------------------------------------------

def fallback_points(df: pd.DataFrame) -> list[dict]:
    """Liefert IMMER >=4 Kandidaten, damit jede Aktie auf 3 Kaufpunkte kommt."""
    last = df.iloc[-1]
    close = float(last["close"])
    out = []
    hi52 = float(last["hi52"])
    out.append({
        "strategie": "Fallback: 52W-Hoch-Breakout",
        "kaufpunkt": round(hi52 * 1.001, 2),
        "stop": round(hi52 * 0.93, 2),
        "ziel": None,
        "status": "Kein Muster — generischer Breakout-Level",
        "notiz": f"52W-Hoch {hi52:.2f}",
    })
    kons_high = float(df["high"].iloc[-20:].max())
    out.append({
        "strategie": "Fallback: 20-Tage-Hoch (Pivot)",
        "kaufpunkt": round(kons_high + 0.01, 2),
        "stop": round(float(df["low"].iloc[-20:].min()) - 0.01, 2),
        "ziel": None,
        "status": "Kein Muster — Konsolidierungs-Pivot",
        "notiz": "Hoch der letzten 20 Handelstage",
    })
    if not pd.isna(last["ma50"]):
        ma50 = float(last["ma50"])
        if close > ma50:
            out.append({
                "strategie": "Fallback: MA50-Pullback",
                "kaufpunkt": round(ma50 * 1.005, 2),
                "stop": round(ma50 * 0.95, 2),
                "ziel": None,
                "status": "Kein Muster — Rücksetzer-Kauf am MA50",
                "notiz": f"MA50 aktuell {ma50:.2f} (nur bei intaktem Trend nutzen)",
            })
        else:
            out.append({
                "strategie": "Fallback: MA50-Rückeroberung",
                "kaufpunkt": round(ma50 * 1.005, 2),
                "stop": round(ma50 * 0.94, 2),
                "ziel": None,
                "status": "Kurs UNTER MA50 — erst bei Reclaim interessant",
                "notiz": f"MA50 aktuell {ma50:.2f}; Kauf erst wenn Schlusskurs drüber",
            })
    hi63 = float(df["high"].iloc[-63:].max())
    out.append({
        "strategie": "Fallback: Quartals-Hoch (63 Tage)",
        "kaufpunkt": round(hi63 + 0.01, 2),
        "stop": round(hi63 * 0.92, 2),
        "ziel": None,
        "status": "Kein Muster — mittelfristiger Widerstand",
        "notiz": "Hoch der letzten 63 Handelstage",
    })
    return out


# ---------------------------------------------------------------------------
# Auswertung je Ticker
# ---------------------------------------------------------------------------

# Die Wochenfassung steht direkt hinter der Tagesfassung: Sie beschreibt
# dasselbe Muster, nur ueber eine laengere Formation. Ohne diesen Eintrag
# landete sie ganz hinten und fiele bei drei Kaufpunkten je Aktie oft
# heraus (Gerhards Ergaenzung vom 04.08.2026).
PRIORITY = ["High & Tight Flag", "HTF Innen-Einstieg", "VCP",
            "Cup & Handle", "Cup & Handle (Wochenbasis)", "Darvas Box",
            "Earnings-Pullback", "Rectangle Top"]


# ---------------------------------------------------------------------------
# Kapitel 9 und 10 — Gerhards Uebergabe vom 02.08.2026
# ---------------------------------------------------------------------------

SHAKEOUT_DATEI = "shakeout_warteliste.json"
FOKUSLISTE_DATEI = "fokusliste.json"


def wiener_zeit():
    """Zeitstempel in Wiener Zeit, mit der Zone im Text.

    Der Nachtscan laeuft auf einem GitHub-Rechner, und der steht auf
    UTC. Ein Lauf um 18:00 New York wird dort als 22:00 des VORTAGES
    geschrieben, obwohl er in Wien um 00:00 desselben Tages stattfand.
    Am 04.08.2026 hat genau das Mathias stutzig gemacht: "Der
    Nachtscanner ist also gestern gelaufen?" - er war es nicht.
    Deshalb steht hier ab jetzt die Zeit, in der er denkt, samt Zone."""
    try:
        from zoneinfo import ZoneInfo
        jetzt = datetime.now(ZoneInfo("Europe/Vienna"))
        return jetzt.strftime("%Y-%m-%d %H:%M") + " Wien"
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M") + " (Zone unbekannt)"


def _json_lesen(pfad, vorgabe):
    try:
        with open(pfad, encoding="utf-8-sig") as f:
            return json.load(f)
    except (OSError, ValueError):
        return vorgabe


def _json_schreiben(pfad, inhalt):
    with open(pfad, "w", encoding="utf-8") as f:
        json.dump(inhalt, f, ensure_ascii=False, indent=1)


def shakeout_durchgang(loaded: dict) -> list[dict]:
    """Kapitel 10 als zehnte Musterpruefung, einmal taeglich nach Schluss.

    Anders als die anderen Musterdetektoren liefert diese Pruefung nicht
    sofort ein Ergebnis: Ein Spring wandert erst auf eine Warteliste und
    wird zum Kaufsignal, wenn ihn spaeter ein Ruecksetzer mit geringerem
    Volumen bestaetigt. Das kann bis zu 15 Handelstage dauern, deshalb
    ueberlebt die Liste in einer eigenen Datei.

    ACHTUNG beim Aufraeumen (Gerhard ausdruecklich): Diese Datei gehoert
    in denselben woechentlichen Rhythmus wie die anderen Zustandsdateien,
    aber es duerfen NUR abgelaufene Eintraege weg — wer die ganze Liste
    leert, verliert die aktiven Wartepositionen."""
    warteliste = _json_lesen(SHAKEOUT_DATEI, {})
    vorher = len(warteliste)
    # Ohne Listenplatz keine Warteposition (24.08.2026, siehe
    # shakeout.warteliste_bereinigen): sonst frieren Eintraege von
    # Aktien, die von der Wochenliste gefallen sind, fuer immer ein.
    gelistet = {t for t, _ in listen.alle_ticker()}
    warteliste, fremde = shakeout.warteliste_bereinigen(warteliste, gelistet)
    if fremde:
        print(f"Shakeout: {len(fremde)} Warteposition(en) von Aktien "
              f"ohne aktuellen Listenplatz entfernt: {', '.join(fremde)}")
    signale = []
    for ticker, (df, company) in loaded.items():
        try:
            kurse = shakeout.aus_scanner_df(df)
            warteliste, signal = shakeout.warte_auf_sekundaertest_und_alarmiere(
                kurse, warteliste, ticker)
        except Exception as e:
            print(f"    Shakeout-Fehler bei {ticker}: {type(e).__name__}: {e}")
            continue
        if signal:
            signal["firma"] = company
            signale.append(signal)
    _json_schreiben(SHAKEOUT_DATEI, warteliste)
    print(f"Shakeout: {len(warteliste)} Aktie(n) warten auf ihren "
          f"Sekundaertest (vorher {vorher}), {len(signale)} bestaetigt.")
    return signale


def crash_support_durchgang(loaded: dict, api_key, limiter) -> list[dict]:
    """Kapitel 8, REKONSTRUIERT — laeuft nur zur Beobachtung mit.

    Die Funde gehen ins Trigger-Logbuch und auf den Bildschirm, aber
    NICHT in die Excel-Mappe und in keine Push-Meldung. Grund steht im
    Kopf von crash_support.py: Von Kapitel 8 gibt es nur fuenf Zahlen
    und einen Satz zum Stopp; Ausloeser, Kaufpunkt und Ziel sind
    abgeleitet. Solange Gerhard das nicht bestaetigt hat, wird danach
    nicht gehandelt.

    Die Strategie ruht ausserdem meistens: Sie ist nur scharf, wenn der
    SPY mindestens 10 % unter seinem 52-Wochen-Hoch steht. Gemessen an
    fuenf Jahren waren das 13 % aller Handelstage."""
    import crash_support

    # Der Scanner führt die Spalten klein geschrieben, crash_support
    # erwartet Yahoos Schreibweise — übersetzt wird an genau der Stelle,
    # die Kapitel 10 dafür schon hat.
    roh_index = fetch_history(crash_support.INDEX, api_key, limiter)
    if roh_index is None:
        print("Kapitel 8: kein Indexkurs — übersprungen.")
        return []
    df_index = shakeout.aus_scanner_df(roh_index)
    scharf, rueck = crash_support.regime_scharf(df_index)
    if rueck is None:
        print("Kapitel 8: kein Indexkurs — übersprungen.")
        return []
    if not scharf:
        print(f"Kapitel 8 (Crash-Support): ruht — {crash_support.INDEX} steht "
              f"{rueck*100:+.1f} % unter dem 52-Wochen-Hoch, scharf ab "
              f"{ZENTRAL['crash_support']['regime_index_drawdown']*100:.0f} %.")
        return []

    kennzahlen = crash_support.hole_kennzahlen(list(loaded))
    signale = []
    for ticker, (df, company) in loaded.items():
        try:
            s = crash_support.pruefe(shakeout.aus_scanner_df(df),
                                     kennzahlen.get(ticker, {}), df_index,
                                     ticker=ticker, firma=company)
        except Exception as e:
            print(f"    Kapitel-8-Fehler bei {ticker}: {type(e).__name__}: {e}")
            continue
        if s:
            signale.append(s)

    if signale:
        trigger_logbuch.protokolliere_viele(signale, quelle="scanner/kapitel8")
    print(f"Kapitel 8 (Crash-Support): scharf ({rueck*100:+.1f} %), "
          f"{len(signale)} Beobachtung(en) — nur ins Logbuch, nicht in die Mappe.")
    return signale


def exit_durchgang(loaded: dict) -> list[dict]:
    """Das Exit-Regelwerk einmal taeglich gegen alle offenen Positionen.

    Hier und nicht im Waechter, aus zwei Gruenden: Die Regeln pruefen
    ausdruecklich auf SCHLUSSKURS-Basis (ein Docht darunter loest nicht
    aus), und der Nachtscan hat die Tagesdaten ohnehin schon geholt.

    Der Zaehler der Handelstage kommt aus der Kurshistorie, nicht aus
    dem Kalender: Die Acht-Wochen-Regel zaehlt HANDELStage, und
    Feiertage wuerden die Frist sonst verkuerzen."""
    bestand = positionen.laden()
    offen = [e for e in bestand.values() if e.get("status") == "offen"]
    if not offen:
        return []

    kurse, ma21, ma50 = {}, {}, {}
    heute_index = 0
    # SEIT KAPITEL 12 (28.08.2026): Beobachtungen heissen 'TICKER|Zusatz',
    # der Kurs haengt am Symbol-Feld. Gebraucht werden die Kurse aller
    # offenen SYMBOLE, egal wie der Eintrag verschluesselt ist.
    gesucht = {e.get("symbol", k) for k, e in bestand.items()
               if e.get("status") == "offen"}
    for ticker, (df, _) in loaded.items():
        if ticker not in gesucht:
            continue
        kurse[ticker] = float(df["close"].iloc[-1])
        if len(df) >= 21:
            ma21[ticker] = float(df["close"].tail(21).mean())
        if len(df) >= 50:
            ma50[ticker] = float(df["close"].tail(50).mean())
        heute_index = max(heute_index, len(df))

    # Einstiegs-Index nachtragen, wo er noch fehlt: Beim Eroeffnen ueber
    # die Befehlszeile ist er null, hier steht die Historie zur Verfuegung.
    for ticker, e in bestand.items():
        if e.get("status") == "offen" and not e.get("einstieg_index"):
            df = loaded.get(e.get("symbol", ticker), (None, None))[0]
            if df is not None and "datetime" in df.columns:
                vorher = df[df["datetime"].astype(str) <= e["einstieg_datum"]]
                e["einstieg_index"] = int(len(vorher))

    meldungen = positionen.pruefe_bestand(
        bestand, kurse, heute_index, ma21=ma21, ma50=ma50)
    positionen.speichern(bestand)
    noch_offen = sum(1 for e in bestand.values() if e.get("status") == "offen")
    print(f"Exit-Regelwerk: {len(offen)} offene Position(en) geprueft, "
          f"{len(meldungen)} Meldung(en), {noch_offen} bleiben offen.")
    return meldungen


def fokusliste_schreiben(loaded: dict) -> int:
    """Kapitel 9: die Bedingungen, die schon am Vorabend feststehen.

    Der Waechter kann das morgens nicht selbst rechnen — fuer das
    RS-Rating braucht es 252 Handelstage Historie und das ganze
    Universum auf einmal. Also wird die Liste hier vorbereitet; am
    Morgen kommt nur noch der Gap der Aktie dazu.

    Gerhards Einschraenkung dazu: Gegen die eigene Kernliste gerechnet
    ist das RS-Rating eine Naeherung. Belastbar wird es erst mit einem
    marktweiten Universum."""
    schluss_je_ticker = {t: df["close"].tolist() for t, (df, _) in loaded.items()}
    ratings = red_to_green.rs_ratings_fuer_universum(schluss_je_ticker)

    eintraege = {}
    for ticker, (df, company) in loaded.items():
        pruef = red_to_green.fokuslisten_kandidat(
            df["close"].tolist(), df["low"].tolist(), ratings.get(ticker))
        if not pruef["ok"]:
            continue
        eintraege[ticker] = {
            "firma": company,
            "vortagesschluss": round(float(df["close"].iloc[-1]), 4),
            # Der Massstab fuer die Volumen-Signatur, hier einmal
            # gerechnet statt am Morgen je Aktie.
            "v50": round(float(df["volume"].tail(50).mean()), 1),
            "rs_rating": pruef["rs_rating"],
            "ueber_52w_tief_pct": pruef["ueber_52w_tief_pct"],
        }
    _json_schreiben(FOKUSLISTE_DATEI, {
        "gebaut_am": wiener_zeit(),
        "universum": len(loaded), "aktien": eintraege,
    })
    _r2g = ZENTRAL["red_to_green"]
    print(f"Fokusliste fuer Red-to-Green: {len(eintraege)} von "
          f"{len(loaded)} Aktien (RS ueber {_r2g['rs_min']}, ueber EMA21 "
          f"und EMA50, mindestens {_r2g['min_ueber_tief']*100:.0f} % "
          f"ueber dem 52-Wochen-Tief).")
    return len(eintraege)


def analyze(df: pd.DataFrame, rs_percentile: float | None,
            darvas_erlaubt: bool = True, ticker: str = "") -> dict:
    """Alle Muster auf eine Aktie.

    darvas_erlaubt (Gerhard, 14.08.2026): Die Darvas Box laeuft
    AUSSCHLIESSLICH auf der Darvas-Liste. Steht die Aktie nur in der
    grossen Liste, wird der Detektor gar nicht erst aufgerufen — nicht
    erst sein Ergebnis verworfen, denn ein nicht gelaufener Detektor
    kann auch keinen Kaufpunkt in die Mappe schreiben. Die Regel selbst
    steht in listen.py und nur dort."""
    df = add_indicators(df)
    last = df.iloc[-1]
    tt_pass, tt_count, tt_failed = check_trend_template(df, rs_percentile)

    hits = []
    detektoren = [detect_htf, detect_htf_innen,
                  lambda d: detect_vcp(d, tt_pass),
                  detect_cup_handle, detect_rectangle,
                  cup_handle_v2.detect_cup_handle_v2]
    if darvas_erlaubt:
        detektoren.insert(3, detect_darvas)
    if ticker:
        # EARNINGS-PULLBACK (Gerhards Freigabe 31.08.2026): braucht als
        # einziger Detektor den Ticker, weil er die Zahlen-Bindung des
        # Gap-Tags belegt (Termin-Historie); die Kursreihe allein kann
        # einen Zahlen-Gap nicht von einem anlasslosen unterscheiden.
        import earnings_pullback
        detektoren.append(
            lambda d: earnings_pullback.detect_earnings_pullback(d, ticker))
    for fn in detektoren:
        try:
            res = fn(df)
        except Exception as e:
            res = None
            print(f"    Detektor-Fehler ({fn}): {e}")
        if res:
            hits.append(res)

    # BEIDE CUP-FASSUNGEN LAUFEN NEBENEINANDER (Gerhard, 04.08.2026), weil
    # die Wochenfassung bewusst gelockerte Toleranzen hat: Auf kurze Cups
    # angewandt liesse sie Fehlsignale durch, die die strengere
    # Tagesfassung zu Recht aussiebt. Feuern ausnahmsweise beide auf
    # derselben Aktie, gewinnt die hoehere Punktzahl — zwei Eintraege fuer
    # dasselbe Muster waeren nur Doppelmeldung.
    cups = [h for h in hits if h["strategie"].startswith("Cup & Handle")]
    if len(cups) > 1:
        bester = max(cups, key=lambda h: h.get("score", 0))
        for c in cups:
            if c is not bester:
                hits.remove(c)

    hits.sort(key=lambda h: PRIORITY.index(h["strategie"]) if h["strategie"] in PRIORITY else 99)
    points = hits[:3]
    if len(points) < 3:
        fbs = fallback_points(df)
        # 1. Pass: nur Levels, die sich von vorhandenen unterscheiden (>0,5 %)
        for fb in fbs:
            if len(points) >= 3:
                break
            if all(abs(fb["kaufpunkt"] - p["kaufpunkt"]) / fb["kaufpunkt"] > 0.005 for p in points):
                points.append(fb)
        # 2. Pass: notfalls trotzdem auffüllen (andere Strategie-Logik, ähnlicher Preis)
        for fb in fbs:
            if len(points) >= 3:
                break
            if all(fb["strategie"] != p["strategie"] for p in points):
                points.append(fb)

    # ZWEI MUSTER AUF DEMSELBEN KAUFPUNKT bekamen bisher zwei
    # VERSCHIEDENE Stops (Gerhard, 06.08.2026, an ANRO aufgefallen: Cup &
    # Handle wollte 23,17, die Darvas Box 25,65 — bei identischem
    # Kaufpunkt 28,86). Die Detektoren kennen einander nicht, jeder
    # rechnet seinen eigenen Bruchpunkt.
    # Der ENGERE Stop gewinnt, dieselbe Regel, nach der der Waechter
    # gleiche Preise zusammenlegt. Der Kaufpunkt bleibt zweimal stehen,
    # damit sichtbar bleibt, welche Muster erfuellt sind; zu zwei
    # MELDUNGEN fuehrt das nicht, das verhindert der Waechter seit dem
    # 28.07.2026 (gemessen: genau diese 7 Faelle legt er zusammen).
    nach_preis = {}
    for p in points:
        if p.get("stop") is None or p.get("kaufpunkt") is None:
            continue
        schluessel = round(p["kaufpunkt"], 2)
        nach_preis.setdefault(schluessel, []).append(p)
    for gleiche in nach_preis.values():
        if len(gleiche) < 2:
            continue
        engster = max(g["stop"] for g in gleiche)
        for g in gleiche:
            if g["stop"] != engster:
                g["stop_eigen"] = g["stop"]      # was dieses Muster wollte
                g["stop"] = engster

    # DER ZEHN-PROZENT-DECKEL, an der einen Stelle, durch die JEDER
    # Kaufpunkt muss — Muster wie Fallbacks (Gerhard, 06.08.2026).
    # Die Regel stand immer im Regelwerk: stop = max(strukturpunkt,
    # kaufpunkt * 0,90). Nur rief sie niemand auf. Jeder Detektor gab
    # seinen Strukturpunkt roh zurueck, und der wanderte unveraendert in
    # die Mappe: 344 von 1098 Paaren lagen ueber dem Deckel (31 %), der
    # weiteste bei 73,2 % (WLFC). Ein Stop 73 % unter dem Kaufpunkt ist
    # kein Stop.
    for p in points:
        exit_regeln.deckel_anwenden(p)

    return {
        "close": float(last["close"]),
        "hi52": float(last["hi52"]),
        "lo52": float(last["lo52"]),
        "rs": rs_percentile,
        "tt_pass": tt_pass,
        "tt_count": tt_count,
        "tt_failed": tt_failed,
        "pattern_count": len(hits),
        "points": points,
    }


# ---------------------------------------------------------------------------
# Excel-Output (farbcodiert)
# ---------------------------------------------------------------------------

def write_excel(rows: list[dict], out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    GREEN = PatternFill("solid", start_color="C6EFCE")
    YELLOW = PatternFill("solid", start_color="FFEB9C")
    GREY = PatternFill("solid", start_color="EDEDED")
    HEAD = PatternFill("solid", start_color="1F4E78")
    thin = Border(*[Side(style="thin", color="BBBBBB")] * 4)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kaufpunkte"
    headers = ["Ticker", "Firma", "Kurs", "52W-Hoch", "52W-Tief", "Abst. 52W-Hoch",
               "RS-Rank", "Trend Template", "Umsatzwachstum", "Gewinnwachstum",
               "KP1 Strategie", "KP1 Preis", "KP1 Abst.", "KP1 Stop", "KP1 Ziel", "KP1 Status",
               "KP2 Strategie", "KP2 Preis", "KP2 Abst.", "KP2 Stop", "KP2 Ziel", "KP2 Status",
               "KP3 Strategie", "KP3 Preis", "KP3 Abst.", "KP3 Stop", "KP3 Ziel", "KP3 Status",
               "Notizen"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def dist(kp, close):
        return f"{(kp / close - 1) * 100:+.1f}%" if kp and close else ""

    # Sortierung: echte Muster zuerst, dann Trend-Template-Treffer
    rows_sorted = sorted(rows, key=lambda r: (-r["res"]["pattern_count"],
                                              not r["res"]["tt_pass"],
                                              r["ticker"]))
    for row in rows_sorted:
        r = row["res"]
        line = [row["ticker"], row["company"], round(r["close"], 2),
                round(r["hi52"], 2), round(r["lo52"], 2),
                f"{(r['close'] / r['hi52'] - 1) * 100:+.1f}%",
                round(r["rs"]) if r["rs"] is not None else "n/a",
                f"✓ 8/8" if r["tt_pass"] else f"✗ {r['tt_count']}/8"]

        # Fundamentaldaten laut Regelwerk. Fehlt der Wert, steht "n/a" —
        # das ist etwas anderes als "Wachstum zu gering" und darf nicht
        # verwechselt werden.
        fund = row.get("fundamentals") or {}
        for schluessel, grenze in (("umsatzwachstum", CFG["umsatz_min_wachstum"]),
                                   ("gewinnwachstum", None)):
            wert = fund.get(schluessel)
            if wert is None:
                line.append("n/a")
            else:
                marke = ""
                if grenze is not None:
                    marke = "✓ " if wert >= grenze else "✗ "
                line.append(f"{marke}{wert * 100:+.0f}%")
        notes = []
        for i in range(3):
            if i < len(r["points"]):
                p = r["points"][i]
                line += [p["strategie"], p["kaufpunkt"], dist(p["kaufpunkt"], r["close"]),
                         p["stop"], p["ziel"] if p["ziel"] else "", p["status"]]
                notes.append(f"KP{i+1}: {p['notiz']}")
            else:
                line += [""] * 6
        if not r["tt_pass"] and r["tt_failed"]:
            notes.append("TT fehlt: " + "; ".join(r["tt_failed"][:3]))
        line.append(" | ".join(notes))
        ws.append(line)

        fill = GREEN if r["pattern_count"] >= 1 else (YELLOW if r["tt_pass"] else GREY)
        for c in ws[ws.max_row]:
            c.fill = fill
            c.border = thin

    widths = [8, 26, 9, 10, 10, 12, 8, 12, 15, 15] + [18, 9, 9, 9, 9, 30] * 3 + [60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "C2"

    # Legende
    lg = wb.create_sheet("Legende")
    lg.append(["Farbe", "Bedeutung"])
    lg.append(["Grün", "Mindestens 1 echtes Chartmuster aktiv (Kaufpunkte = Muster-Trigger)"])
    lg.append(["Gelb", "Kein Muster, aber Minervini Trend Template 8/8 erfüllt (Fallback-Level)"])
    lg.append(["Grau", "Weder Muster noch Trend Template — Fallback-Level nur zur Orientierung"])
    lg.append([])
    lg.append(["Hinweis", "Alle Breakout-Kaufpunkte gelten nur mit Volumen-Bestätigung "
                          "(Regelwerk). RS-Rank = Perzentil innerhalb der gescannten Liste."])
    lg["A1"].fill = HEAD; lg["B1"].fill = HEAD
    lg["A1"].font = Font(bold=True, color="FFFFFF"); lg["B1"].font = Font(bold=True, color="FFFFFF")
    lg["A2"].fill = GREEN; lg["A3"].fill = YELLOW; lg["A4"].fill = GREY
    lg.column_dimensions["A"].width = 12; lg.column_dimensions["B"].width = 95

    wb.save(out_path)


# ---------------------------------------------------------------------------
# ntfy-Push (optional)
# ---------------------------------------------------------------------------

def push_ntfy(topic: str, rows: list[dict]):
    hot = [r for r in rows if r["res"]["pattern_count"] >= 1]
    if not hot:
        return
    lines = []
    for r in hot[:15]:
        p = r["res"]["points"][0]
        lines.append(f"{r['ticker']}: {p['strategie']} — KP {p['kaufpunkt']}")
    body = "\n".join(lines)
    try:
        r = requests.post(f"https://ntfy.sh/{topic}", data=body.encode("utf-8"),
                          headers={"Title": f"Pattern-Scanner: {len(hot)} Treffer"}, timeout=15)
        # Kennung merken, damit der Freitags-Putz auch diese Meldung
        # wieder wegraeumen kann (siehe ntfy_verlauf.py).
        ntfy_verlauf.merke_antwort(r)
        print(f"Push an ntfy.sh/{topic} gesendet ({len(hot)} Treffer).")
    except Exception as e:
        print(f"ntfy-Fehler: {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def load_tickers(csv_path: str) -> list[tuple[str, str]]:
    """Eine einzelne CSV. Fuer BEIDE Listen siehe listen.alle_ticker()."""
    df = pd.read_csv(csv_path)
    tcol = next((c for c in df.columns if c.strip().lower() == "ticker"), None)
    if tcol is None:
        sys.exit("CSV enthält keine 'Ticker'-Spalte.")
    ccol = next((c for c in df.columns if c.strip().lower() == "company"), None)
    out, seen = [], set()
    for _, row in df.iterrows():
        t = str(row[tcol]).strip().upper()
        if t and t != "NAN" and t not in seen:
            seen.add(t)
            out.append((t, str(row[ccol]) if ccol else ""))
    return out


def main():
    ap = argparse.ArgumentParser(description="Pattern-Scanner nach Regelwerk (6 Muster)")
    ap.add_argument("csv", help="Die grosse Finviz-CSV mit Ticker-Spalte")
    ap.add_argument("--darvas", default=None,
                    help="Die Darvas-Liste (Vorgabe: darvas.csv, wenn "
                         "vorhanden). Nur auf ihr entstehen Darvas-Kaufpunkte.")
    ap.add_argument("--out", default="kaufpunkte.xlsx", help="Excel-Ausgabedatei")
    ap.add_argument("--rate", type=int, default=8, help="API-Calls pro Minute (Free-Tier: 8)")
    ap.add_argument("--ntfy", default=None, help="ntfy.sh-Topic für Push (optional)")
    ap.add_argument("--limit", type=int, default=None, help="Nur die ersten N Ticker (zum Testen)")
    args = ap.parse_args()

    # Der Schluessel ist nur noch fuer die Rueckfallebene noetig: Hauptquelle
    # ist der Yahoo-Sammelabruf. Ohne Schluessel laeuft alles weiter, solange
    # Yahoo antwortet — faellt Yahoo aus, fehlen dann allerdings die Daten.
    api_key = os.environ.get("TWELVE_DATA_API_KEY")
    if not api_key:
        print("⚠ Kein TWELVE_DATA_API_KEY gesetzt — es gibt dann keine "
              "Rückfallebene, falls Yahoo ausfällt.")

    # ZWEI LISTEN (Gerhard, 14.08.2026): Gescannt wird die VEREINIGUNG
    # beider, damit alle Muster ueberall laufen; nur die Darvas Box bleibt
    # auf die Darvas-Liste beschraenkt. Die Regel steht in listen.py.
    import listen
    darvas_pfad = args.darvas or listen.DARVAS_DATEI
    tickers = listen.alle_ticker(haupt=args.csv, darvas=darvas_pfad)
    darvas_erlaubte = {t.upper() for t, _ in listen.darvas_liste(darvas_pfad)}
    print(listen.uebersicht(haupt=args.csv, darvas=darvas_pfad))
    # MARKTAMPEL (Gerhards Freigabe 31.08.2026, Baustein 4): einmal je
    # Nachtscan berechnen, BEVOR die erste Logbuch-Zeile entsteht —
    # trigger_logbuch haengt die Farbe an jede Zeile. Die Ampel
    # informiert nur; gefiltert wird nichts, erst die Auswertung
    # "Trefferquote je Farbe" entscheidet, ob sie je mehr darf.
    try:
        import marktampel
        _ampel = marktampel.aktualisieren()
        if _ampel:
            print(f"Marktampel: {_ampel['farbe'].upper()} "
                  f"(Handelstag {_ampel['handelstag']})")
    except Exception as e:
        print(f"Marktampel nicht berechenbar: {type(e).__name__}: {e}")
    hinweis = listen.fehlende_liste(haupt=args.csv, darvas=darvas_pfad)
    if hinweis:
        # NICHT nur ins Protokoll: Der Hinweis wandert unten auch in die
        # Mappe, weil das Protokoll niemand liest (Mathias, 14.08.2026).
        print(f"  ACHTUNG: {hinweis}")
    if args.limit:
        tickers = tickers[: args.limit]
    if not tickers:
        sys.exit("Keine Ticker in den Listen gefunden.")
    dauer = len(tickers) / args.rate
    print(f"{len(tickers)} Ticker geladen. Bei {args.rate} Calls/min dauert das "
          f"~{dauer:.0f} Minuten (Cache-Treffer sind gratis).")
    if len(tickers) > 750:
        print("⚠ ACHTUNG: Twelve-Data-Free-Tier erlaubt nur 800 API-Calls pro TAG. "
              f"Bei {len(tickers)} Tickern wird das Limit gerissen — Aktien am Ende der "
              "Liste liefern dann Fehler. Optionen: Liste splitten und an 2 Tagen laufen "
              "lassen (Cache merkt sich Tag 1), oder Twelve-Data-Bezahlplan.")
    if dauer > 170:
        print("⚠ Hinweis: Läuft das in GitHub Actions, muss timeout-minutes im Workflow "
              f"über {dauer:.0f} liegen (Maximum bei GitHub: 360).")

    limiter = RateLimiter(args.rate)

    # Zuerst der Sammelabruf: holt alles auf einmal und macht die Schleife
    # unten praktisch kostenlos. Faellt er aus, geht jeder Ticker einzeln
    # ueber Twelve Data — langsamer, aber es laeuft.
    alle_ticker = [t for t, _ in tickers] + [BENCHMARK]
    lade_yahoo_sammelabruf(alle_ticker)

    # 1. Durchlauf: Daten holen + RS-Rohscore
    loaded, raw_rs = {}, {}
    for i, (ticker, company) in enumerate(tickers, 1):
        print(f"[{i}/{len(tickers)}] {ticker} …")
        df = fetch_history(ticker, api_key, limiter)
        if df is None:
            continue
        loaded[ticker] = (df, company)
        s = rs_score(df)
        if s is not None:
            raw_rs[ticker] = s

    # RS-Perzentile innerhalb der Liste
    rs_pct = {}
    if raw_rs:
        ser = pd.Series(raw_rs)
        rs_pct = (ser.rank(pct=True) * 100).to_dict()

    # Fundamentaldaten laut Regelwerk (Umsatzwachstum-Filter)
    fundamentals = hole_fundamentals(list(loaded.keys()))

    # 2. Durchlauf: Muster analysieren
    rows = []
    for ticker, (df, company) in loaded.items():
        res = analyze(df, rs_pct.get(ticker),
                      darvas_erlaubt=ticker.upper() in darvas_erlaubte,
                      ticker=ticker)
        # KELL-ZYKLUS (31.08.2026, nur Messung): die Phase je Aktie,
        # fuers Logbuch — siehe kell_zyklus.py.
        try:
            import kell_zyklus
            res["zyklus"] = kell_zyklus.klassifiziere(df)
        except Exception:
            res["zyklus"] = None
        rows.append({"ticker": ticker, "company": company, "res": res,
                     "fundamentals": fundamentals.get(ticker, {})})
        tag = "🟢" if res["pattern_count"] else ("🟡" if res["tt_pass"] else "⚪")
        pats = ", ".join(p["strategie"] for p in res["points"] if not p["strategie"].startswith("Fallback"))
        print(f"  {tag} {ticker}: {res['pattern_count']} Muster"
              + (f" ({pats})" if pats else ""))

    # Kapitel 10: die zehnte Musterpruefung. Laeuft NACH den anderen,
    # weil sie eine eigene Warteliste ueber Tage hinweg fuehrt und nicht
    # in die Excel-Mappe gehoert.
    shakeout_signale = shakeout_durchgang(loaded)

    # Kapitel 9: die Fokusliste fuer den Live-Waechter von morgen.
    fokusliste_schreiben(loaded)

    # VOLUMENKURVEN, eine je Aktie (Gerhard, 06.08.2026). Sie gehören
    # hierher und nicht in den Wächter: Seine Vorgabe lautet "einmal pro
    # Tag und Aktie, dann zwischenspeichern", und der Nachtlauf ist genau
    # dieser eine Zeitpunkt. Der Wächter zur Eröffnung hätte sonst
    # hunderte Abrufe zu erledigen, während jede Sekunde zählt.
    try:
        volumen.baue_kurven(sorted({t for t, _ in tickers}))
    except Exception as e:
        print(f"  ⚠ Volumenkurven konnten nicht gebaut werden "
              f"({type(e).__name__}: {e}) — die betroffenen Aktien gelten "
              f"morgen als NICHT VERIFIZIERBAR. Geschätzt wird nichts.")

    # ZAHLEN-TERMINE (Gerhard, 12.08.2026). Wie die Volumenkurven einmal
    # am Tag hier und nicht im Waechter: Ein Abruf je Aktie kostet Zeit,
    # und der Waechter braucht sie zur Eroeffnung nicht.
    try:
        zahlen_termine.baue(sorted({t for t, _ in tickers}), leise=False)
    except Exception as e:
        print(f"  ⚠ Zahlen-Termine konnten nicht geholt werden "
              f"({type(e).__name__}: {e}) — es wird nichts vermerkt. "
              f"Ein fehlender Vermerk ist KEIN Beweis, dass keine kommen.")

    # Kapitel 8, rekonstruiert: laeuft nur mit, meldet nur ins Logbuch.
    crash_support_durchgang(loaded, api_key, limiter)

    # Das Exit-Regelwerk gegen die offenen Positionen (Gerhard, 05.08.2026).
    exit_meldungen = exit_durchgang(loaded)

    # Alles Gefundene mitschreiben — auch was nicht gekauft wird. Ohne
    # das bleibt jedes "per Mitschreiben verfeinern" aus den Uebergaben
    # ein frommer Wunsch, weil die Signale nach der Meldung weg sind.
    trigger_logbuch.protokolliere_viele(
        [{"ticker": s["symbol"], "firma": s.get("firma", ""),
          "strategie": "Shakeout-Spring", "kaufpunkt": s["kaufpunkt"],
          "stop": s["stop"], "ziel": s["kursziel"],
          "volumen_typ": s.get("volumen_typ", "")} for s in shakeout_signale],
        quelle="scanner/kapitel10")
    # Ein Eintrag je MUSTER, nicht je Aktie: Jedes Muster hat seinen
    # eigenen Kaufpunkt und Stopp, und genau die will man spaeter
    # gegeneinander messen koennen. Fallback-Punkte bleiben draussen,
    # sie sind kein Signal.
    trigger_logbuch.protokolliere_viele(
        [{"ticker": r["ticker"], "firma": r.get("company", ""),
          "strategie": p["strategie"], "kaufpunkt": p.get("kaufpunkt"),
          "stop": p.get("stop"), "ziel": p.get("ziel"),
          "status": p.get("status", ""),
          "rs": r["res"].get("rs"), "trend_template": r["res"].get("tt_pass"),
          "zyklus": r["res"].get("zyklus"),
          "schluss": r["res"].get("close")}
         for r in rows if r["res"]["pattern_count"] >= 1
         for p in r["res"]["points"]
         if not p["strategie"].startswith("Fallback")],
        quelle="scanner/nachtlauf")

    write_excel(rows, args.out)
    print(f"\nFertig → {args.out}")

    # Die Exit-Meldungen zuerst: Was zu VERKAUFEN ist, ist dringlicher als
    # der naechste Kaufpunkt.
    if exit_meldungen:
        print("\nExit-Regelwerk, fällige Handlungen:")
        for m in exit_meldungen:
            print("  " + positionen.melde_text(m))

    if shakeout_signale:
        print("\nShakeout-Spring, Sekundaertest bestaetigt:")
        for s in shakeout_signale:
            print(f"  {s['symbol']} ({s.get('firma','')}): Kaufpunkt "
                  f"{s['kaufpunkt']}, Stop {s['stop']}, Ziel {s['kursziel']}; "
                  f"{s['volumen_typ']}")
        # Kapitel 12: Shakeout-Signale entstehen HIER, nicht im Waechter —
        # ihre Beobachtungen werden deshalb hier eroeffnet.
        try:
            gewinnzonen_lauf.beobachtungen_aus_shakeout(shakeout_signale)
        except Exception as e:
            print(f"  Beobachtungs-Fuetterung (Shakeout) fehlgeschlagen: "
                  f"{type(e).__name__}: {e}")

    # KAPITEL 12 (Gerhards Uebergabe vom 28.08.2026): Gewinnzonen je
    # offener Beobachtung. Laeuft NACH write_excel, damit der
    # Stop-Nachzug die frischen Strukturpunkte dieser Nacht sieht.
    # Gemeldet wird nichts um Mitternacht — die Befunde landen in
    # exit_befunde.json, der Waechter meldet sie zur Handelszeit
    # (dasselbe Muster wie beim Sektor-Radar). Auch die
    # Kapitel-11-Meldungen oben gehen seither diesen Weg statt nur ins
    # Protokoll.
    try:
        gewinnzonen_lauf.gewinn_durchgang(loaded, args.out,
                                          exit_meldungen=exit_meldungen)
    except Exception as e:
        print(f"Kapitel 12 fehlgeschlagen: {type(e).__name__}: {e}")
    n_green = sum(1 for r in rows if r["res"]["pattern_count"] >= 1)
    n_tt = sum(1 for r in rows if r["res"]["tt_pass"])
    print(f"Treffer: {n_green} mit aktivem Muster, {n_tt} bestehen das Trend Template.")

    if args.ntfy:
        push_ntfy(args.ntfy, rows)


if __name__ == "__main__":
    main()
